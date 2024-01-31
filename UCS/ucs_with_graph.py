import openpyxl
import matplotlib.pyplot as plt
import time
import tracemalloc

excel_file_path = './map16x22.xlsx'

class BasicNode:
    def __init__(self, rowNumber, cellNumber, value, cost):
        self.rowNumber = rowNumber
        self.cellNumber = cellNumber
        self.cost = cost
        self.value = value
        self.next_nodes = []
        self.next_locations = []

class Node:
    visited = []
    nodeHolder = {}
    keyHolder = []
    finalPath = []
    finalCost = 0

    def nodePlanner(row, cell, value):
        key = f"{row}:{cell}"
        
        try:    
            node_value, cost = value.split(',')
        except:
            node_value = value
            cost = 0
        nodeValue = BasicNode(row, cell, node_value, cost)
        nodeValue.next_locations = [f'{row-1}:{cell}',f'{row+1}:{cell}',f'{row}:{cell-1}',f'{row}:{cell+1}']
        
        Node.nodeHolder[key] = nodeValue
        
        return nodeValue

    def nodeChecking(rowNumber, cellNumber):
        checker = f"{rowNumber}:{cellNumber}"
        if checker in Node.keyHolder:
            retrieve = Node.nodeHolder[checker]
            return retrieve
        return None

    def nodeAppending():
        for key, value in Node.nodeHolder.items():
            for locations in Node.nodeHolder[key].next_locations:
                if locations in Node.nodeHolder:
                    Node.nodeHolder[key].next_nodes.append(Node.nodeHolder[locations])

    def nodeCreation(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        startNode = None; goalNode = None
        for index, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start = 1):
            for col_num, value in enumerate(row, start=1):
                if value == 0:
                    continue
                node = Node.nodePlanner(index, col_num, value)
                
                if value == 'start':
                    startNode = node
                if value == 'goal':
                    goalNode = node
              
        return startNode, goalNode


    def get_obstacle_locations(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        obstacles = [(row, col) for row in range(1, 23) for col in range(1, 17) if sheet.cell(row=row, column=col).value == 0]
        return obstacles

    def UCS_Traversal(current, goal, path=[], cost=0):
        Node.visited.append(current)

        if current.value == goal.value:
            print("Goal Found")
            Node.finalPath = path + [current]
            Node.finalCost = cost
            return True
        else:
            i = 0; j = 0
            while i < len(current.next_nodes):
                j = 0
                while j < len(current.next_nodes):
                    if int(current.next_nodes[i].cost) < int(current.next_nodes[j].cost):
                        temp = current.next_nodes[j]
                        current.next_nodes[j] = current.next_nodes[i]
                        current.next_nodes[i] = temp
                    j = j + 1
                i = i + 1

            for nodes in current.next_nodes:
                if nodes in Node.visited:
                    continue

                #Visualization using matplotlib
                # plt.clf()  # Clear the current figure

                # # Plot obstacles as black lines
                # obstacle_locations = Node.get_obstacle_locations(excel_file_path)
                # for obstacle in obstacle_locations:
                #     plt.plot([obstacle[1] - 1], [obstacle[0] - 1], marker='o', color='black', markersize=8)

                # # Plot the path as a blue line
                # for idx in range(len(path) - 1):
                #     plt.plot([path[idx].cellNumber - 1, path[idx + 1].cellNumber - 1],
                #              [path[idx].rowNumber - 1, path[idx + 1].rowNumber - 1],
                #              color='blue', linestyle='-', linewidth=2)

                # # Plot other elements with different colors and markers
                # plt.scatter([current.cellNumber - 1], [current.rowNumber - 1], color='yellow', marker='o', s=50, label='Current Node')  # Current Node
                # plt.scatter([goal.cellNumber - 1], [goal.rowNumber - 1], color='limegreen', marker='s', s=50, label='Goal Node')  # Goal Node

                # # Adjust plot settings
                # #plt.axis('equal')  # Set equal scaling for x and y axes
                # #plt.xticks([])  # Hide x-axis ticks
                # #plt.yticks([])  # Hide y-axis ticks
                # plt.pause(0.5)

                cost += int(current.cost)
                final = Node.UCS_Traversal(nodes, goal, path + [current], cost)
                if final:
                    return True
            return False

    def UCS_Call(start, goal):
        start_time = time.time()        # Starting Time Tracking
        tracemalloc.start()             # Starting Memory Tracking
        Node.UCS_Traversal(start, goal)
        end_time = time.time()
        total_time = end_time - start_time
        print(f"Total Execution Time : {total_time}")
        if Node.finalPath:
            finalPlot = []
            for path in Node.finalPath:
                print(f"{path.rowNumber, path.cellNumber}")
                finalPlot.append((path.rowNumber - 1, path.cellNumber - 1))
            print(f"Final Cost is {Node.finalCost}")

            # Visualization using matplotlib
            # plt.clf()  

            # obstacle_locations = Node.get_obstacle_locations(excel_file_path)
            # for obstacle in obstacle_locations:
            #     plt.plot([obstacle[1] - 1], [obstacle[0] - 1], color='black', marker='o', markersize=10, linewidth=5)

            # for idx in range(len(Node.finalPath) - 1):
            #     plt.plot([Node.finalPath[idx].cellNumber - 1, Node.finalPath[idx + 1].cellNumber - 1],
            #              [Node.finalPath[idx].rowNumber - 1, Node.finalPath[idx + 1].rowNumber - 1],
            #              color='blue', linestyle='dashed', linewidth=2)

            # plt.scatter([start.cellNumber - 1], [start.rowNumber - 1], color='yellow', marker='o', s=20, label='Start Node')  # Start Node
            # plt.scatter([goal.cellNumber - 1], [goal.rowNumber - 1], color='limegreen', marker='s', s=20, label='Goal Node')  # Goal Node
            # plt.scatter([Node.finalPath[-1].cellNumber - 1], [Node.finalPath[-1].rowNumber - 1], color='red', marker='x', s=200, label='End Node')  # End Node

            # plt.title("Maze Visualization")
            # plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
            # plt.show()

            # Printing Memory Usage
            current, peak = tracemalloc.get_traced_memory()
            print(f"Current memory usage: {current / 10**6} MB")
            print(f"Peak memory usage: {peak / 10**6} MB")
            tracemalloc.stop()

# Example usage
start, goal = Node.nodeCreation(excel_file_path)
print(f"Start Value {start.value} and cost is {start.cost}")
Node.nodeAppending()
Node.UCS_Call(start, goal)